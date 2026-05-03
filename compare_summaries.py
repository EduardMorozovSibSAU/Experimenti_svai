"""
compare_summaries.py
Ищет все summary.xlsx в папке Results_* (корень + exp_N подпапки),
сравнивает R2 / RMSE / MAE попарно и сохраняет Excel с цветовой гаммой.
"""

import os
import re
import itertools
from pathlib import Path

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ── настройки ────────────────────────────────────────────────────────────────
ROOT = Path(".")
OUTPUT_FILE = ROOT / "comparison_summaries.xlsx"
RESULTS_PATTERN = "Results_*"   # ищем все папки Results_…
KEY_COL = "Soil_types"          # колонка-ключ для join
METRICS = ["R2", "RMSE", "MAE"] # метрики для сравнения
# ─────────────────────────────────────────────────────────────────────────────


def find_summaries(root: Path) -> dict[str, pd.DataFrame]:
    """Рекурсивно ищет summary.xlsx в Results_* папках. Возвращает {label: df}."""
    found = {}
    for results_dir in sorted(root.glob(RESULTS_PATTERN)):
        if not results_dir.is_dir():
            continue
        # корень папки
        p = results_dir / "summary.xlsx"
        if p.exists():
            label = results_dir.name
            found[label] = pd.read_excel(p)
        # подпапки exp_N
        for sub in sorted(results_dir.iterdir()):
            if sub.is_dir():
                p2 = sub / "summary.xlsx"
                if p2.exists():
                    label = f"{results_dir.name}/{sub.name}"
                    found[label] = pd.read_excel(p2)
    return found


def make_diff_df(df_a: pd.DataFrame, df_b: pd.DataFrame,
                 label_a: str, label_b: str) -> pd.DataFrame:
    """
    Джойнит два summary по KEY_COL и вычисляет diff = B - A для каждой метрики.
    Для R2: положительный diff = B лучше.
    Для RMSE/MAE: отрицательный diff = B лучше.
    """
    a = df_a[[KEY_COL] + METRICS].copy()
    b = df_b[[KEY_COL] + METRICS].copy()
    merged = a.merge(b, on=KEY_COL, suffixes=(f"__{label_a}", f"__{label_b}"))

    rows = []
    for _, row in merged.iterrows():
        entry = {KEY_COL: row[KEY_COL]}
        for m in METRICS:
            va = row[f"{m}__{label_a}"]
            vb = row[f"{m}__{label_b}"]
            diff = vb - va
            entry[f"{m}_{label_a}"] = va
            entry[f"{m}_{label_b}"] = vb
            entry[f"{m}_diff(B-A)"] = diff
        rows.append(entry)
    return pd.DataFrame(rows)


# ── цвета ────────────────────────────────────────────────────────────────────
GREEN_STRONG = PatternFill("solid", fgColor="63BE7B")
GREEN_LIGHT  = PatternFill("solid", fgColor="A9D18E")
NEUTRAL      = PatternFill("solid", fgColor="FFEB84")
RED_LIGHT    = PatternFill("solid", fgColor="F4A460")
RED_STRONG   = PatternFill("solid", fgColor="F8696B")
HEADER_FILL  = PatternFill("solid", fgColor="2F75B6")
HEADER_FONT  = Font(color="FFFFFF", bold=True)
SUBHEADER_FILL = PatternFill("solid", fgColor="BDD7EE")

def color_for_diff(metric: str, diff: float, vmax: float) -> PatternFill:
    """Возвращает заливку в зависимости от направления улучшения."""
    if pd.isna(diff) or vmax == 0:
        return NEUTRAL
    ratio = abs(diff) / vmax  # нормируем по максимуму в колонке
    better = (diff > 0) if metric == "R2" else (diff < 0)
    if better:
        return GREEN_STRONG if ratio > 0.3 else GREEN_LIGHT
    else:
        return RED_STRONG if ratio > 0.3 else RED_LIGHT


# Ширина одной мини-таблички: KEY_COL + для каждой метрики (A, B, diff) = 1 + 3*n
BLOCK_COLS = 1 + len(METRICS) * 3   # колонок в одном блоке сравнения
BLOCK_ROWS = None                    # будет определено по данным

DIAG_FILL = PatternFill("solid", fgColor="D9D9D9")  # серый — диагональ


def _write_mini_block(ws, row0: int, col0: int,
                      diff_df: pd.DataFrame, label_a: str, label_b: str,
                      vmax: dict):
    """Рисует мини-блок сравнения A vs B начиная с ячейки (row0, col0)."""
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # строка заголовка блока
    end_col = col0 + BLOCK_COLS - 1
    ws.merge_cells(start_row=row0, start_column=col0,
                   end_row=row0, end_column=end_col)
    hcell = ws.cell(row0, col0, f"B={label_b}  diff=B−A")
    hcell.fill = HEADER_FILL
    hcell.font = HEADER_FONT
    hcell.alignment = Alignment(horizontal="center")

    # строка подзаголовков метрик
    ws.cell(row0+1, col0, KEY_COL).fill = SUBHEADER_FILL
    ws.cell(row0+1, col0).font = Font(bold=True)
    c = col0 + 1
    for m in METRICS:
        ws.merge_cells(start_row=row0+1, start_column=c,
                       end_row=row0+1, end_column=c+2)
        cell = ws.cell(row0+1, c, m)
        cell.fill = SUBHEADER_FILL
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        c += 3

    # строка подзаголовков A/B/diff
    ws.cell(row0+2, col0, "").fill = SUBHEADER_FILL
    c = col0 + 1
    for m in METRICS:
        for lbl in [f"{m}(A)", f"{m}(B)", "diff"]:
            cell = ws.cell(row0+2, c, lbl)
            cell.fill = SUBHEADER_FILL
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
            c += 1

    # данные
    for i, (_, row) in enumerate(diff_df.iterrows()):
        er = row0 + 3 + i
        ws.cell(er, col0, row[KEY_COL]).border = border
        c = col0 + 1
        for m in METRICS:
            va   = row.get(f"{m}_{label_a}", np.nan)
            vb   = row.get(f"{m}_{label_b}", np.nan)
            diff = row.get(f"{m}_diff(B-A)", np.nan)
            for val, fmt, is_diff in [
                (va,   "0.0000",           False),
                (vb,   "0.0000",           False),
                (diff, "+0.0000;-0.0000",  True),
            ]:
                cell = ws.cell(er, c)
                cell.value = round(float(val), 4) if not pd.isna(val) else None
                cell.number_format = fmt
                cell.border = border
                cell.alignment = Alignment(horizontal="right")
                if is_diff:
                    cell.fill = color_for_diff(m, diff, vmax[m])
                c += 1


def _write_diag_block(ws, row0: int, col0: int, label: str, n_rows: int):
    """Серый блок на диагонали."""
    end_col = col0 + BLOCK_COLS - 1
    end_row = row0 + 2 + n_rows  # заголовок(2) + данные
    ws.merge_cells(start_row=row0, start_column=col0,
                   end_row=end_row, end_column=end_col)
    cell = ws.cell(row0, col0, label)
    cell.fill = DIAG_FILL
    cell.font = Font(bold=True, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_matrix_sheet(writer, summaries: dict[str, pd.DataFrame]):
    """Один лист MATRIX: шахматная матрица всех попарных сравнений."""
    wb = writer.book
    ws = wb.create_sheet(title="MATRIX", index=0)

    labels = list(summaries.keys())
    n = len(labels)
    n_data = max(len(df) for df in summaries.values())  # строк данных в блоке
    block_height = 3 + n_data  # заголовок(3) + данные

    # Предвычислим vmax глобально по всем diff
    all_diffs = {m: [] for m in METRICS}
    for i in range(n):
        for j in range(n):
            if i == j:
                continue
            diff_df = make_diff_df(summaries[labels[i]], summaries[labels[j]],
                                   labels[i], labels[j])
            for m in METRICS:
                col_name = f"{m}_diff(B-A)"
                if col_name in diff_df.columns:
                    all_diffs[m].extend(diff_df[col_name].abs().dropna().tolist())
    vmax = {m: max(vals) if vals else 1 for m, vals in all_diffs.items()}

    # Заголовки строк (левый столбец) и столбцов (верхняя строка)
    LABEL_COL_W = 18   # ширина левого служебного столбца
    GAP = 1            # зазор между блоками (строки/колонки)

    # Рассчитываем смещения блоков
    # col_offset[j] = стартовая колонка блока j (1-based), учитываем служебный col=1
    col_offsets = []
    cur_col = 2
    for j in range(n):
        col_offsets.append(cur_col)
        cur_col += BLOCK_COLS + GAP

    row_offsets = []
    cur_row = 3  # строки 1-2 — заголовки столбцов
    for i in range(n):
        row_offsets.append(cur_row)
        cur_row += block_height + GAP

    thin = Side(style="thin", color="888888")

    # Заголовки столбцов (строки 1-2)
    for j, label in enumerate(labels):
        c0 = col_offsets[j]
        ws.merge_cells(start_row=1, start_column=c0,
                       end_row=2, end_column=c0 + BLOCK_COLS - 1)
        cell = ws.cell(1, c0, label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Заголовки строк (колонка 1)
    for i, label in enumerate(labels):
        r0 = row_offsets[i]
        ws.merge_cells(start_row=r0, start_column=1,
                       end_row=r0 + block_height - 1, end_column=1)
        cell = ws.cell(r0, 1, label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True, text_rotation=90)

    # Блоки
    for i in range(n):
        for j in range(n):
            r0 = row_offsets[i]
            c0 = col_offsets[j]
            if i == j:
                _write_diag_block(ws, r0, c0, labels[i], n_data)
            else:
                diff_df = make_diff_df(
                    summaries[labels[i]], summaries[labels[j]],
                    labels[i], labels[j]
                )
                _write_mini_block(ws, r0, c0, diff_df, labels[i], labels[j], vmax)

    # Ширина колонок
    ws.column_dimensions["A"].width = LABEL_COL_W
    for j in range(n):
        c0 = col_offsets[j]
        ws.column_dimensions[get_column_letter(c0)].width = 20     # KEY_COL
        for k in range(1, BLOCK_COLS):
            ws.column_dimensions[get_column_letter(c0 + k)].width = 9

    # Высота строк-заголовков
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20


def write_combined_sheet(writer, all_summaries: dict[str, pd.DataFrame]):
    """Лист со всеми summary рядом (без diff) для общего обзора."""
    wb = writer.book
    ws = wb.create_sheet(title="ALL_summaries", index=0)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    col_offset = 1
    for label, df in all_summaries.items():
        sub = df[[KEY_COL] + METRICS].copy()
        n_cols = len(sub.columns)

        # метка-заголовок
        ws.merge_cells(start_row=1, start_column=col_offset,
                       end_row=1, end_column=col_offset + n_cols - 1)
        hcell = ws.cell(1, col_offset, label)
        hcell.fill = HEADER_FILL
        hcell.font = HEADER_FONT
        hcell.alignment = Alignment(horizontal="center")

        # заголовки колонок
        for c_idx, col_name in enumerate(sub.columns, start=col_offset):
            cell = ws.cell(2, c_idx, col_name)
            cell.fill = SUBHEADER_FILL
            cell.font = Font(bold=True)
            cell.border = border

        # данные
        for r_idx, row in sub.iterrows():
            for c_idx, val in enumerate(row, start=col_offset):
                cell = ws.cell(r_idx + 3, c_idx)
                if isinstance(val, float):
                    cell.value = round(val, 4)
                    cell.number_format = "0.0000"
                else:
                    cell.value = val
                cell.border = border

        # ширина
        ws.column_dimensions[get_column_letter(col_offset)].width = 20
        for i in range(col_offset + 1, col_offset + n_cols):
            ws.column_dimensions[get_column_letter(i)].width = 10

        col_offset += n_cols + 1  # пробел между блоками


# ── main ─────────────────────────────────────────────────────────────────────

def main():
    summaries = find_summaries(ROOT)
    if len(summaries) < 2:
        print(f"Найдено summary: {len(summaries)}. Нужно минимум 2 для сравнения.")
        for k in summaries:
            print(f"  {k}")
        return

    print(f"Найдено {len(summaries)} summary:")
    for k in summaries:
        print(f"  {k}  ({len(summaries[k])} кластеров)")

    labels = list(summaries.keys())
    print(f"Пар для сравнения: {len(list(itertools.combinations(labels, 2)))}")

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        # создаём пустую книгу (ExcelWriter требует хотя бы один лист)
        writer.book.create_sheet("_tmp")

        write_matrix_sheet(writer, summaries)
        write_combined_sheet(writer, summaries)

        # удаляем временный лист
        del writer.book["_tmp"]

    print(f"\nГотово → {OUTPUT_FILE.resolve()}")


if __name__ == "__main__":
    main()
